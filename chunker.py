import os
import json
from datetime import datetime, date
from decimal import Decimal
import numpy as np
from openpyxl import load_workbook


INPUT_EXCEL_DIR = r"C:\Users\GANNOJU SHAHSANK\Downloads\MAANG_PYTHON\GenAI\Financial_rag_bot\rag_json_approach\Data"
OUTPUT_CHUNKS_DIR = r"C:\Users\GANNOJU SHAHSANK\Downloads\MAANG_PYTHON\GenAI\Financial_rag_bot\rag_json_approach\chunks\previous_chunks"


# =========================
# NORMALIZATION
# =========================
def normalize_value(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    return v


# =========================
# BASIC HELPERS
# =========================
def is_empty_row(row):
    return all(c in (None, "", " ") for c in row)


def is_numeric(v):
    return isinstance(v, (int, float, Decimal, np.integer, np.floating))


def count_numeric(row):
    return sum(1 for c in row if is_numeric(c))


def row_to_text(row):
    return " | ".join(str(c).strip() for c in row if c not in (None, "", " "))


# =========================
# ROW TYPE DETECTION
# =========================
def looks_like_fund_row(row):
    """
    Fund rows:
    - first cell is long text
    - contains NO digits in name
    - may have date / NAV later
    """
    first = row[0]
    if isinstance(first, str):
        if len(first) > 12 and not any(ch.isdigit() for ch in first):
            return True
    return False


def is_data_row(row):
    # must have numeric AND must NOT be fund row
    return count_numeric(row) >= 1 and not looks_like_fund_row(row)


def is_column_header_row(row):
    """
    Column headers:
    - >=2 non-empty cells
    - ZERO numeric cells
    - contiguous
    """
    non_empty = [v for v in row if v not in (None, "", " ")]
    if len(non_empty) < 2:
        return False

    if any(is_numeric(v) for v in non_empty):
        return False

    idxs = [i for i, v in enumerate(row) if v not in (None, "", " ")]
    return max(idxs) - min(idxs) <= len(idxs) + 2


# =========================
# COLUMN MERGE
# =========================
def merge_column_headers(header_rows):
    max_cols = max(len(r) for r in header_rows)
    merged = []
    for c in range(max_cols):
        parts = []
        for r in header_rows:
            if c < len(r) and r[c] not in (None, "", " "):
                parts.append(str(r[c]).strip())
        merged.append(" | ".join(parts) if parts else None)
    return merged


# =========================
# CORE
# =========================
def process_excel_file(excel_path):
    file_name = os.path.splitext(os.path.basename(excel_path))[0]
    out_dir = os.path.join(OUTPUT_CHUNKS_DIR, file_name)
    os.makedirs(out_dir, exist_ok=True)

    wb = load_workbook(excel_path, data_only=True)
    chunks_written = 0

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]

        # -------- GLOBAL HEADER (FIRST 2 ROWS ONLY)
        global_header = []
        for i in (0, 1):
            if i < len(rows):
                txt = row_to_text(rows[i])
                if txt:
                    global_header.append(txt)

        # -------- HARD TABLE SPLIT ON EMPTY ROW
        tables = []
        current = []

        for excel_row, r in enumerate(rows[2:], start=3):
            if is_empty_row(r):
                if current:
                    tables.append(current)
                    current = []
                continue
            current.append((excel_row, r))

        if current:
            tables.append(current)

        # -------- PROCESS EACH TABLE
        for table in tables:
            # find column headers ANYWHERE in table
            col_header_rows = [r for _, r in table if is_column_header_row(r)]
            if not col_header_rows:
                continue

            columns = merge_column_headers(col_header_rows)

            # find first column header index
            first_header_idx = min(
                i for i, (_, r) in enumerate(table) if is_column_header_row(r)
            )

            # -------- SUBHEADERS = 1–2 ROWS ABOVE FIRST COLUMN HEADER
            subheaders = []
            start = max(0, first_header_idx - 2)
            for _, r in table[start:first_header_idx]:
                if count_numeric(r) <= 1:
                    txt = row_to_text(r)
                    if txt:
                        subheaders.append(txt)

            # -------- DATA ROWS = BELOW COLUMN HEADERS
            for excel_row, r in table[first_header_idx + 1:]:
                if not is_data_row(r):
                    continue

                chunk = {
                    "source_file": file_name,
                    "sheet_name": sheet_name,
                    "excel_row_number": excel_row,
                    "global_header": global_header,
                    "subheaders": subheaders,
                    "data": {}
                }

                for i, v in enumerate(r):
                    if i < len(columns) and columns[i] and v is not None:
                        chunk["data"][columns[i]] = normalize_value(v)

                if chunk["data"]:
                    with open(
                        os.path.join(out_dir, f"{sheet_name}_row_{excel_row}.json"),
                        "w",
                        encoding="utf-8"
                    ) as f:
                        json.dump(chunk, f, indent=2)
                    chunks_written += 1

    if chunks_written == 0:
        os.rmdir(out_dir)


# =========================
# RUNNER
# =========================
def main():
    os.makedirs(OUTPUT_CHUNKS_DIR, exist_ok=True)
    for f in os.listdir(INPUT_EXCEL_DIR):
        if f.lower().endswith(".xlsx"):
            process_excel_file(os.path.join(INPUT_EXCEL_DIR, f))
    print("✅ DONE — no bleed, empty rows hard-stop, fund rows excluded")


if __name__ == "__main__":
    main()
